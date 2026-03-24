"""
SafariFlow Flask API v5
Added: JWT approval token generation, /approve and /reject endpoints
Fixed: confirm_payment agent_id mismatch
Fixed: markup loop saves cost_unit_price, cost_total_price, markup_pct, profit
"""

import os
import io
import json
import uuid
import base64
import hmac
import hashlib
import logging
import urllib.request
import urllib.error
import urllib.parse
import time
from flask import Flask, request, jsonify, redirect, send_file
from pdf_generator import generate_quote_pdf

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ─── CORS ─────────────────────────────────────────────────────────────────────
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, PATCH, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, apikey'
    return response

@app.route('/', defaults={'path': ''}, methods=['OPTIONS'])
@app.route('/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    return '', 204

# ─── Environment variables ────────────────────────────────────────────────────
SUPABASE_URL        = os.environ.get('SUPABASE_URL', '').rstrip('/')
SUPABASE_KEY        = os.environ.get('SUPABASE_SERVICE_KEY', '')
ANTHROPIC_API_KEY   = os.environ.get('ANTHROPIC_API_KEY', '')
APPROVAL_SECRET     = os.environ.get('APPROVAL_SECRET', 'safariflow-secret-change-me')
MAKE_S2_WEBHOOK     = os.environ.get('MAKE_SCENARIO2_WEBHOOK', '')
MAKE_S3_WEBHOOK     = os.environ.get('MAKE_SCENARIO3_WEBHOOK', '')
API_BASE_URL        = os.environ.get('API_BASE_URL', 'https://web-production-4788f.up.railway.app')
PORTAL_URL          = os.environ.get('PORTAL_URL', 'https://safariflow-portal.netlify.app')
RESEND_API_KEY      = os.environ.get('RESEND_API_KEY', '')
RESEND_FROM         = os.environ.get('RESEND_FROM', 'SafariFlow <onboarding@resend.dev>')
UNSPLASH_ACCESS_KEY = os.environ.get('UNSPLASH_ACCESS_KEY', '')
CLERK_WEBHOOK_SECRET = os.environ.get('CLERK_WEBHOOK_SECRET', '')
STORAGE_BUCKET      = 'quote-pdfs'
OUTPUT_DIR          = os.path.join(os.path.dirname(__file__), 'outputs')
TOKEN_EXPIRY_DAYS   = 7

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ─── JWT-style token helpers ──────────────────────────────────────────────────
def generate_token(quote_id, action):
    expires = int(time.time()) + (TOKEN_EXPIRY_DAYS * 86400)
    payload = f"{quote_id}:{action}:{expires}"
    sig = hmac.new(
        APPROVAL_SECRET.encode(),
        payload.encode(),
        hashlib.sha256
    ).hexdigest()[:16]
    token_str = f"{payload}:{sig}"
    return base64.urlsafe_b64encode(token_str.encode()).decode()


def verify_token(token, expected_action):
    try:
        decoded = base64.urlsafe_b64decode(token.encode()).decode()
        parts = decoded.split(':')
        if len(parts) != 4:
            return None
        quote_id, action, expires, sig = parts
        if action != expected_action:
            return None
        if int(expires) < int(time.time()):
            return None
        expected_payload = f"{quote_id}:{action}:{expires}"
        expected_sig = hmac.new(
            APPROVAL_SECRET.encode(),
            expected_payload.encode(),
            hashlib.sha256
        ).hexdigest()[:16]
        if not hmac.compare_digest(sig, expected_sig):
            return None
        return quote_id
    except Exception:
        return None


# ─── Supabase helpers ─────────────────────────────────────────────────────────
def supabase_get(table, params=None):
    if not SUPABASE_URL or not SUPABASE_KEY:
        return []
    try:
        query = f"{SUPABASE_URL}/rest/v1/{table}"
        if params:
            query += '?' + urllib.parse.urlencode(params)
        req = urllib.request.Request(query, headers={
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'apikey': SUPABASE_KEY,
            'Content-Type': 'application/json',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        logger.error(f"Supabase fetch error ({table}): {str(e)}")
        return []


def supabase_update(table, match_params, update_data):
    if not SUPABASE_URL or not SUPABASE_KEY:
        logger.error("Supabase credentials missing — cannot update")
        return False
    try:
        query = f"{SUPABASE_URL}/rest/v1/{table}?" + urllib.parse.urlencode(match_params)
        payload = json.dumps(update_data).encode('utf-8')
        logger.info(f"PATCH URL: {query}")
        logger.info(f"PATCH body: {update_data}")
        req = urllib.request.Request(query, data=payload, method='PATCH', headers={
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'apikey': SUPABASE_KEY,
            'Content-Type': 'application/json',
            'Prefer': 'return=representation',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = resp.read().decode()
            logger.info(f"PATCH response: {result}")
        return True
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        logger.error(f"Supabase PATCH HTTP error: {e.code} {e.reason} — {error_body}")
        return False
    except Exception as e:
        logger.error(f"Supabase PATCH error: {str(e)}")
        return False


def supabase_upload(file_path, filename):
    if not SUPABASE_URL or not SUPABASE_KEY:
        logger.error("Supabase credentials missing — cannot upload")
        return ''
    try:
        upload_url = f"{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{filename}"
        logger.info(f"Uploading to: {upload_url}")
        with open(file_path, 'rb') as f:
            pdf_bytes = f.read()
        logger.info(f"File size: {len(pdf_bytes)} bytes")
        req = urllib.request.Request(upload_url, data=pdf_bytes, method='POST', headers={
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/pdf',
            'x-upsert': 'true',
        })
        with urllib.request.urlopen(req, timeout=120) as resp:
            response_body = resp.read().decode()
            logger.info(f"Upload response: {response_body}")
        public_url = f"{SUPABASE_URL}/storage/v1/object/public/{STORAGE_BUCKET}/{filename}"
        logger.info(f"Uploaded to Supabase Storage: {public_url}")
        return public_url
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        logger.error(f"Supabase upload HTTP error: {e.code} {e.reason} — {error_body}")
        return ''
    except Exception as e:
        logger.error(f"Supabase upload error: {str(e)}")
        return ''


def trigger_make_webhook(webhook_url, payload):
    if not webhook_url:
        logger.warning("Make webhook URL not set")
        return False
    try:
        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(webhook_url, data=data, method='POST', headers={
            'Content-Type': 'application/json',
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            resp.read()
        logger.info(f"Make webhook triggered: {webhook_url}")
        return True
    except Exception as e:
        logger.error(f"Make webhook error: {str(e)}")
        return False


# ─── Email helper ─────────────────────────────────────────────────────────────
BREVO_API_KEY  = os.environ.get('BREVO_API_KEY', '')
GMAIL_USER     = os.environ.get('GMAIL_USER', '')
GMAIL_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD', '')
logger.info(f"Email config: Brevo={'YES' if BREVO_API_KEY else 'NO'}, Resend={'YES' if os.environ.get('RESEND_API_KEY') else 'NO'}")

def send_email(to, subject, html, attachments=None):
    """Send email — Brevo API primary, Resend fallback."""
    to_list = [to] if isinstance(to, str) else to

    # ── Brevo API ─────────────────────────────────────────────────────────────
    if BREVO_API_KEY:
        try:
            payload = {
                'sender':      {'name': 'SafariFlow', 'email': 'ephraim063@gmail.com'},
                'to':          [{'email': t} for t in to_list],
                'subject':     subject,
                'htmlContent': html,
            }
            if attachments:
                payload['attachment'] = [
                    {'name': att['filename'], 'content': att['content']}
                    for att in attachments
                ]
            data = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(
                'https://api.brevo.com/v3/smtp/email',
                data=data, method='POST',
                headers={
                    'api-key':      BREVO_API_KEY,
                    'Content-Type': 'application/json',
                    'Accept':       'application/json',
                }
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                result = json.loads(resp.read().decode())
                logger.info(f"Email sent via Brevo: {result.get('messageId', 'ok')}")
            return True
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            logger.error(f"Brevo error: {e.code} — {error_body}")
        except Exception as e:
            logger.error(f"Brevo error: {str(e)}")

    # ── Resend fallback ───────────────────────────────────────────────────────
    if not RESEND_API_KEY:
        logger.warning("No email provider configured — skipping email")
        return False
    try:
        payload = {
            'from': RESEND_FROM,
            'to': to_list,
            'subject': subject,
            'html': html,
        }
        if attachments:
            payload['attachments'] = attachments
        data = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(
            'https://api.resend.com/emails',
            data=data, method='POST',
            headers={
                'Authorization': f'Bearer {RESEND_API_KEY}',
                'Content-Type': 'application/json',
            }
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode())
            logger.info(f"Email sent via Resend: {result.get('id')}")
        return True
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        logger.error(f"Resend error: {e.code} error code: {error_body}")
        return False
    except Exception as e:
        logger.error(f"Resend error: {str(e)}")
        return False


def agent_approval_email_html(agent_name, agency_name, client_name, quote_number,
                               start_date, end_date, total_price,
                               approve_url, reject_url,
                               brand_primary='#2E4A7A', brand_secondary='#C4922A'):
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f4f4f4;">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:30px 0;">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.1);">
<tr><td style="background-color:{brand_primary};padding:30px;text-align:center;">
  <h1 style="margin:0;color:#FFFFFF;font-size:22px;letter-spacing:2px;">{agency_name}</h1>
  <p style="margin:6px 0 0;color:#FFFFFF;font-size:12px;letter-spacing:1px;">NEW QUOTE READY FOR REVIEW</p>
</td></tr>
<tr><td style="background-color:{brand_secondary};height:3px;"></td></tr>
<tr><td style="padding:36px 40px;background:#ffffff;">
  <p style="color:#1A1A1A;font-size:16px;margin:0 0 8px;">Hello <strong>{agent_name}</strong>, a new safari quote has been generated and is ready for your review.</p>
  <table width="100%" cellpadding="12" style="background:#F5F0E8;border-radius:6px;margin:16px 0;">
    <tr>
      <td style="color:#444444;font-size:11px;letter-spacing:1px;font-weight:bold;">CLIENT</td>
      <td style="color:#444444;font-size:11px;letter-spacing:1px;font-weight:bold;">QUOTE NUMBER</td>
    </tr>
    <tr>
      <td style="color:#1A1A1A;font-size:14px;font-weight:bold;">{client_name}</td>
      <td style="color:#1A1A1A;font-size:14px;font-weight:bold;">{quote_number}</td>
    </tr>
    <tr>
      <td style="color:#444444;font-size:11px;letter-spacing:1px;font-weight:bold;padding-top:8px;">TRAVEL DATES</td>
      <td style="color:#444444;font-size:11px;letter-spacing:1px;font-weight:bold;padding-top:8px;">TOTAL VALUE</td>
    </tr>
    <tr>
      <td style="color:#1A1A1A;font-size:14px;font-weight:bold;">{start_date} — {end_date}</td>
      <td style="color:#1A1A1A;font-size:14px;font-weight:bold;">${total_price:,.0f}</td>
    </tr>
  </table>
  <table width="100%" cellpadding="0" cellspacing="0" style="margin:16px 0 8px;">
    <tr><td align="center">
      <a href="{approve_url}" style="display:inline-block;background-color:{brand_primary};color:#FFFFFF;padding:16px 32px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:14px;letter-spacing:1px;">
        ✅ APPROVE &amp; SEND TO CLIENT
      </a>
    </td></tr>
  </table>
  <table width="100%" cellpadding="0" cellspacing="0" style="margin:8px 0 16px;">
    <tr><td align="center">
      <a href="{reject_url}" style="display:inline-block;background-color:{brand_secondary};color:#FFFFFF;padding:16px 32px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:14px;letter-spacing:1px;">
        ✏️ MAKE CHANGES
      </a>
    </td></tr>
  </table>
  <table width="100%" cellpadding="10" style="background:#FFF5F5;border-radius:6px;border:1px solid #F5CCCC;margin-bottom:16px;">
    <tr><td style="color:#B03030;font-size:13px;font-weight:bold;text-align:center;">⚠ THESE LINKS WILL EXPIRE IN 7 DAYS</td></tr>
  </table>
  <p style="color:#1A1A1A;font-size:14px;margin:0;">This quote was generated automatically by SafariFlow.</p>
</td></tr>
<tr><td style="background:#F5F0E8;padding:20px 40px;text-align:center;">
  <p style="margin:0;color:#444444;font-size:12px;">{agency_name} · Powered by SafariFlow</p>
</td></tr>
</table>
</td></tr>
</table>
</body></html>"""


def client_quote_email_html(client_name, agency_name, agent_name, agent_email, agent_phone,
                             quote_number, start_date, end_date,
                             accept_url, changes_url,
                             brand_primary='#2E4A7A', brand_secondary='#C4922A'):
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f4f4f4;">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:30px 0;">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.1);">
<tr><td style="background-color:{brand_primary};padding:30px;text-align:center;">
  <h1 style="margin:0;color:#FFFFFF;font-size:22px;letter-spacing:2px;">{agency_name}</h1>
  <p style="margin:6px 0 0;color:#FFFFFF;font-size:12px;letter-spacing:1px;">TRAVEL &amp; SAFARI SPECIALISTS</p>
</td></tr>
<tr><td style="background-color:{brand_secondary};height:3px;"></td></tr>
<tr><td style="padding:36px 40px;background:#ffffff;">
  <p style="color:#1A1A1A;font-size:16px;margin:0 0 8px;">Dear <strong>{client_name}</strong>, your personalised safari quote is ready for review.</p>
  <table width="100%" cellpadding="12" style="background:#F5F0E8;border-radius:6px;margin:16px 0;">
    <tr>
      <td style="color:#444444;font-size:11px;letter-spacing:1px;font-weight:bold;">QUOTE NUMBER</td>
      <td style="color:#444444;font-size:11px;letter-spacing:1px;font-weight:bold;">TRAVEL DATES</td>
    </tr>
    <tr>
      <td style="color:#1A1A1A;font-size:14px;font-weight:bold;">{quote_number}</td>
      <td style="color:#1A1A1A;font-size:14px;font-weight:bold;">{start_date} — {end_date}</td>
    </tr>
  </table>
  <p style="color:#1A1A1A;font-size:15px;text-align:center;">Please review the attached PDF and let us know your decision.</p>
  <table width="100%" cellpadding="0" cellspacing="0" style="margin:16px 0 8px;">
    <tr><td align="center">
      <a href="{accept_url}" style="display:inline-block;background-color:{brand_primary};color:#FFFFFF;padding:16px 32px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:14px;letter-spacing:1px;">
        TAP THIS LINK TO ACCEPT THIS QUOTE
      </a>
    </td></tr>
  </table>
  <table width="100%" cellpadding="0" cellspacing="0" style="margin:8px 0 16px;">
    <tr><td align="center">
      <a href="{changes_url}" style="display:inline-block;background-color:{brand_secondary};color:#FFFFFF;padding:16px 32px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:14px;letter-spacing:1px;">
        TAP THIS LINK TO REQUEST CHANGES
      </a>
    </td></tr>
  </table>
  <table width="100%" cellpadding="10" style="background:#FFF5F5;border-radius:6px;border:1px solid #F5CCCC;margin-bottom:16px;">
    <tr><td style="color:#B03030;font-size:13px;font-weight:bold;text-align:center;">⚠ THESE LINKS WILL EXPIRE IN 7 DAYS</td></tr>
  </table>
  <p style="color:#1A1A1A;font-size:15px;">We look forward to crafting your perfect safari experience.</p>
</td></tr>
<tr><td style="background:#F5F0E8;padding:20px 40px;text-align:center;">
  <p style="margin:0;color:#444444;font-size:12px;">{agency_name} · {agent_email} · {agent_phone}</p>
</td></tr>
</table>
</td></tr>
</table>
</body></html>"""


# ─── Claude API helper ────────────────────────────────────────────────────────
def call_claude(prompt, max_tokens=4000):
    if not ANTHROPIC_API_KEY:
        return {}
    try:
        payload = json.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": max_tokens,
            "messages": [{"role": "user", "content": prompt}]
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload, method='POST',
            headers={
                'x-api-key': ANTHROPIC_API_KEY,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            }
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode())
        text = result.get('content', [{}])[0].get('text', '{}').strip()
        if text.startswith('```'):
            lines = [l for l in text.split('\n') if not l.strip().startswith('```')]
            text = '\n'.join(lines).strip()
        return json.loads(text)
    except json.JSONDecodeError as e:
        logger.error(f"Claude JSON parse error: {str(e)}")
        return {}
    except Exception as e:
        logger.error(f"Claude API error: {str(e)}")
        return {}


# ─── Unsplash photo fetcher ───────────────────────────────────────────────────
def fetch_unsplash_photo(query, width=800, height=500):
    if not UNSPLASH_ACCESS_KEY:
        return None
    try:
        safe_query = urllib.parse.quote(query)
        url = f"https://api.unsplash.com/search/photos?query={safe_query}&per_page=1&orientation=landscape&client_id={UNSPLASH_ACCESS_KEY}"
        req = urllib.request.Request(url, headers={'User-Agent': 'SafariFlow/1.0'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode())
        results = data.get('results', [])
        if not results:
            return None
        photo_url = results[0].get('urls', {}).get('regular', '')
        if not photo_url:
            return None
        img_req = urllib.request.Request(photo_url, headers={'User-Agent': 'SafariFlow/1.0'})
        with urllib.request.urlopen(img_req, timeout=10) as img_resp:
            return img_resp.read()
    except Exception as e:
        logger.warning(f"Unsplash fetch failed for '{query}': {str(e)}")
        return None


def fetch_photos_for_itinerary(itinerary):
    if not UNSPLASH_ACCESS_KEY:
        logger.info("Unsplash key not set — skipping photos")
        return {}

    import threading

    photo_cache = {}
    lock = threading.Lock()

    queries = {}
    for day in itinerary:
        query = day.get('image_search_query') or f"{day.get('destination', '')} safari wildlife Kenya"
        day_num = day.get('day_number', 0)
        queries[day_num] = query

    def fetch_one(day_num, query):
        img_bytes = fetch_unsplash_photo(query)
        if img_bytes:
            with lock:
                photo_cache[str(day_num)] = img_bytes
            logger.info(f"Photo fetched for day {day_num}: {query}")
        else:
            logger.warning(f"No photo for day {day_num}: {query}")

    threads = []
    for day_num, query in queries.items():
        t = threading.Thread(target=fetch_one, args=(day_num, query))
        t.start()
        threads.append(t)

    for t in threads:
        t.join(timeout=20)

    logger.info(f"Photos fetched: {len(photo_cache)}/{len(queries)} days")
    return photo_cache


# ─── Clerk Webhook ────────────────────────────────────────────────────────────
@app.route('/clerk-webhook', methods=['POST'])
def clerk_webhook():
    try:
        payload = request.get_data()
        svix_id        = request.headers.get('svix-id', '')
        svix_timestamp = request.headers.get('svix-timestamp', '')
        svix_signature = request.headers.get('svix-signature', '')

        if CLERK_WEBHOOK_SECRET:
            signed_content = f"{svix_id}.{svix_timestamp}.{payload.decode('utf-8')}"
            secret_bytes = base64.b64decode(CLERK_WEBHOOK_SECRET.replace('whsec_', ''))
            expected_sig = 'v1,' + base64.b64encode(
                hmac.new(secret_bytes, signed_content.encode(), hashlib.sha256).digest()
            ).decode()
            sigs = svix_signature.split(' ')
            if not any(s == expected_sig for s in sigs):
                logger.warning("Clerk webhook signature verification failed")
                return jsonify({'error': 'Invalid signature'}), 401

        data = json.loads(payload)
        event_type = data.get('type')
        logger.info(f"Clerk webhook received: {event_type}")

        if event_type == 'user.created':
            user_data   = data.get('data', {})
            clerk_id    = user_data.get('id', '')
            email       = ''
            first_name  = user_data.get('first_name', '')
            last_name   = user_data.get('last_name', '')
            agent_name  = f"{first_name} {last_name}".strip() or 'Safari Agent'

            emails = user_data.get('email_addresses', [])
            if emails:
                email = emails[0].get('email_address', '')

            existing = supabase_get('agents', {
                'clerk_user_id': f'eq.{clerk_id}',
                'select': 'id'
            })
            if existing:
                logger.info(f"Agent already exists for clerk_id: {clerk_id}")
                return jsonify({'status': 'exists'}), 200

            agent_id = str(uuid.uuid4())
            new_agent = {
                'id': agent_id,
                'clerk_user_id': clerk_id,
                'agent_name': agent_name,
                'agency_name': f"{agent_name} Safaris",
                'email': email,
                'brand_color_primary': '#2E4A7A',
                'brand_color_secondary': '#C4922A',
                'deposit_percentage': 30,
                'balance_due_days': 60,
                'markup_type': 'overall',
                'markup_overall_pct': 20,
                'subscription_status': 'trial',
                'subscription_plan': 'free',
            }

            insert_url = f"{SUPABASE_URL}/rest/v1/agents"
            insert_payload = json.dumps(new_agent).encode('utf-8')
            req = urllib.request.Request(
                insert_url,
                data=insert_payload,
                method='POST',
                headers={
                    'Authorization': f'Bearer {SUPABASE_KEY}',
                    'apikey': SUPABASE_KEY,
                    'Content-Type': 'application/json',
                    'Prefer': 'return=representation',
                }
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                resp.read()

            logger.info(f"Agent created: {agent_id} for {email}")

            if email and RESEND_API_KEY:
                welcome_html = f"""
                <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 16px rgba(0,0,0,0.1)">
                  <div style="background:#2E4A7A;padding:32px;text-align:center;color:white">
                    <h1 style="margin:0;font-size:24px;letter-spacing:2px">SAFARIFLOW</h1>
                    <p style="margin:8px 0 0;opacity:0.8;font-size:13px">WHERE SAFARI BUSINESS GROWS</p>
                  </div>
                  <div style="background:#C4922A;height:3px"></div>
                  <div style="padding:32px">
                    <h2 style="color:#1A1A1A;margin-bottom:8px">Welcome, {agent_name}! 🦁</h2>
                    <p style="color:#444;line-height:1.7">Your SafariFlow account is ready.</p>
                    <a href="{PORTAL_URL}" style="display:inline-block;background:#2E4A7A;color:white;padding:14px 28px;border-radius:8px;text-decoration:none;font-weight:bold;letter-spacing:1px">OPEN MY PORTAL →</a>
                  </div>
                </div>"""
                send_email(
                    to=email,
                    subject="Welcome to SafariFlow — Your AI Safari Quoting Platform 🦁",
                    html=welcome_html
                )

        return jsonify({'status': 'ok'}), 200

    except Exception as e:
        logger.error(f"Clerk webhook error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Generate PDF (Quote) ─────────────────────────────────────────────────────
@app.route('/generate-pdf', methods=['POST'])
def generate_pdf():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': 'No JSON body received'}), 400

        agent_id          = data.get('agent_id', '')
        request_id        = data.get('request_id', str(uuid.uuid4())[:8].upper())
        client_name       = data.get('client_name', '')
        client_email      = data.get('client_email', '')
        client_phone      = data.get('client_phone', '')
        client_nationality= data.get('client_nationality', 'international')
        pax_adults        = int(data.get('pax_adults', 2) or 2)
        pax_children      = int(data.get('pax_children', 0) or 0)
        destinations      = data.get('destination', '')
        start_date        = data.get('start_date', '')
        end_date          = data.get('end_date', '')
        duration_days     = int(data.get('duration_days', 7) or 7)
        accommodation_tier= data.get('accommodation_tier', 'luxury')
        budget_usd        = float(data.get('budget_usd', 10000) or 10000)
        special_requests  = data.get('special_requests', '')

        source_raw = data.get('source', 'manual')
        source_map = {
            'portal': 'form', 'tally': 'form', 'form': 'form',
            'email': 'email', 'safaribookings': 'email',
            'chatbot': 'chatbot', 'manual': 'manual',
        }
        intake_source = source_map.get(source_raw, 'manual')

        logger.info(f"Generating quote for {client_name} — {destinations}")

        agents   = supabase_get('agents', {'id': f'eq.{agent_id}', 'select': '*'})
        agent    = agents[0] if agents else {}
        profiles = supabase_get('agent_profiles', {'agent_id': f'eq.{agent_id}', 'select': '*'})
        profile  = profiles[0] if profiles else {}
        reviews  = supabase_get('agent_reviews', {'agent_id': f'eq.{agent_id}', 'select': '*', 'limit': '3'})

        accommodations = supabase_get('accommodations', {'agent_id': f'eq.{agent_id}', 'is_active': 'eq.true', 'select': '*'})
        transport      = supabase_get('transport_routes', {'agent_id': f'eq.{agent_id}', 'is_active': 'eq.true', 'select': '*'})
        park_fees      = supabase_get('park_fees', {'agent_id': f'eq.{agent_id}', 'select': '*'})

        logger.info(f"Found: {len(accommodations)} accommodations, {len(transport)} transport, {len(park_fees)} park fees")

        def cents(val):
            return round((val or 0) / 100, 2)

        accom_for_claude = [{'name': a.get('name'), 'destination': a.get('destination'), 'category': a.get('category'), 'room_type': a.get('room_type'), 'meal_plan': a.get('meal_plan'), 'price_per_person_usd': cents(a.get('price_per_person_usd_cents'))} for a in accommodations]
        transport_for_claude = [{'from': t.get('from_location'), 'to': t.get('to_location'), 'type': t.get('transport_type'), 'operator': t.get('operator_name'), 'price_per_person_usd': cents(t.get('price_per_person_usd_cents')), 'duration_hours': t.get('duration_hours')} for t in transport]
        fees_for_claude = [{'park': f.get('park_name'), 'destination': f.get('destination'), 'visitor_category': f.get('visitor_category'), 'fee_per_person_per_day_usd': cents(f.get('fee_per_person_per_day_usd_cents'))} for f in park_fees]

        claude2_prompt = f"""You are a safari itinerary pricing specialist. Build the optimal itinerary using ONLY the exact pricing data provided below.

Return ONLY a valid JSON object. No explanation. No markdown. No code fences. Start your response with {{ and end with }}.

Structure:
{{
  "itinerary": [{{"day_number": 1, "date": "YYYY-MM-DD", "destination": "string", "title": "string", "accommodation_name": "string", "room_type": "string", "meal_plan": "string", "nights": 1, "transport_description": "string or null", "image_search_query": "string"}}],
  "line_items": [{{"line_type": "accommodation|transport|park_fee", "description": "string", "details": "string", "quantity": 2, "unit_price": 1950.00, "total_price": 3900.00}}],
  "total_price_usd": 0, "deposit_amount_usd": 0, "balance_amount_usd": 0, "within_budget": true, "budget_notes": "string or null"
}}

RULES: Sequence destinations logically. Select best-value accommodation for tier. Include all transport and park fees. deposit_amount_usd = total * 0.30.

TRIP REQUIREMENTS:
{json.dumps({"destinations": destinations, "duration_days": duration_days, "start_date": start_date, "end_date": end_date, "pax_adults": pax_adults, "pax_children": pax_children, "accommodation_tier": accommodation_tier, "budget_usd": budget_usd, "special_requests": special_requests, "visitor_category": "non_resident"})}

AVAILABLE ACCOMMODATIONS: {json.dumps(accom_for_claude)}
AVAILABLE TRANSPORT: {json.dumps(transport_for_claude)}
PARK FEES: {json.dumps(fees_for_claude)}"""

        itinerary_data = call_claude(claude2_prompt, max_tokens=4000)
        itinerary  = itinerary_data.get('itinerary', [])
        line_items = itinerary_data.get('line_items', [])
        total_price= float(itinerary_data.get('total_price_usd', 0) or 0)

        # ── Apply agent markup — saves cost data for agent breakdown ──────────
        markup_type    = agent.get('markup_type', 'overall')
        markup_overall = float(agent.get('markup_overall_pct', 0) or 0) / 100
        markup_map = {
            'accommodation': float(agent.get('markup_accommodation_pct', 0) or 0) / 100,
            'transport':     float(agent.get('markup_transport_pct', 0) or 0) / 100,
            'park_fee':      float(agent.get('markup_park_fees_pct', 0) or 0) / 100,
            'activity':      float(agent.get('markup_activities_pct', 0) or 0) / 100,
        }

        marked_up_items = []
        for item in line_items:
            item = dict(item)
            if markup_type == 'overall':
                pct = markup_overall
            else:
                line_type = item.get('line_type', 'accommodation')
                pct = markup_map.get(line_type, markup_overall)

            # Save pre-markup (cost) prices for agent cost breakdown
            item['cost_unit_price']  = round(float(item.get('unit_price', 0)), 2)
            item['cost_total_price'] = round(float(item.get('total_price', 0)), 2)
            item['markup_pct']       = round(pct * 100, 2)

            if pct > 0:
                item['unit_price']  = round(float(item.get('unit_price', 0)) * (1 + pct), 2)
                item['total_price'] = round(float(item.get('total_price', 0)) * (1 + pct), 2)

            item['profit'] = round(item['total_price'] - item['cost_total_price'], 2)
            marked_up_items.append(item)

        line_items  = marked_up_items
        total_price = sum(float(i.get('total_price', 0)) for i in line_items)
        deposit     = round(total_price * (float(agent.get('deposit_percentage', 30) or 30) / 100), 2)
        balance     = round(total_price - deposit, 2)

        logger.info(f"Itinerary built: {len(itinerary)} days, total ${total_price} (after markup)")

        first_name = client_name.split()[0] if client_name else "Dear Guest"
        intro_narrative = f"{first_name}, your {duration_days}-day safari across {destinations} has been carefully crafted to deliver an authentic East African experience. Every detail has been arranged to ensure your journey is seamless and unforgettable."
        narrative_days = [{'day_number': d.get('day_number'), 'narrative': f"Today you explore {d.get('destination')} with your expert guide, discovering the remarkable wildlife and landscapes that make this destination truly special.", 'highlight': f"Wildlife encounters in {d.get('destination')}", 'accommodation_description': f"{d.get('accommodation_name')}, {d.get('room_type')} — your comfortable base for the night."} for d in itinerary]

        quote_number         = f"QT-{request_id}"
        approve_token        = generate_token(quote_number, 'approve')
        reject_token         = generate_token(quote_number, 'reject')
        client_accept_token  = generate_token(quote_number, 'client-accept')
        client_changes_token = generate_token(quote_number, 'client-changes')
        approve_url          = f"{API_BASE_URL}/approve?token={approve_token}"
        reject_url           = f"{API_BASE_URL}/reject?token={reject_token}"
        client_accept_url    = f"{API_BASE_URL}/client-accept?token={client_accept_token}"
        client_changes_url   = f"{API_BASE_URL}/client-changes?token={client_changes_token}"

        logger.info(f"Approval tokens generated for {quote_number}")

        filename    = f"SafariFlow_Quote_{quote_number}.pdf"
        output_path = os.path.join(OUTPUT_DIR, filename)

        reviews_formatted = [{'review_text': r.get('review_text', ''), 'client_name': r.get('client_name', ''), 'client_origin': r.get('client_origin', ''), 'trip_summary': r.get('trip_summary', '')} for r in reviews]

        logger.info("Fetching destination photos from Unsplash...")
        photo_cache = fetch_photos_for_itinerary(itinerary)

        pdf_data = {
            'quote_id':     quote_number,
            'generated_at': start_date[:10] if start_date else '',
            'accept_url':   '#',
            'changes_url':  '#',
            'inclusions':   '- All accommodation as specified\n- All meals as per itinerary\n- All game drives\n- Park and conservancy fees\n- Internal flights as specified\n- Airport transfers',
            'exclusions':   '- International flights\n- Travel insurance\n- Visa fees\n- Personal expenses\n- Gratuities',
            'terms':        agent.get('cancellation_terms') or 'This quote is valid for 14 days. A 30% deposit is required to confirm the booking. Balance due 60 days prior to departure.',
            'agent':        {'name': agent.get('agent_name', ''), 'email': agent.get('email', ''), 'phone': agent.get('phone', ''), 'agency': agent.get('agency_name', ''), 'logo_url': agent.get('logo_url', ''), 'website': agent.get('website', '')},
            'client':       {'name': client_name, 'email': client_email, 'phone': client_phone, 'pax_adults': str(pax_adults), 'pax_children': str(pax_children), 'nationality': client_nationality},
            'trip':         {'title': f"{duration_days}-Day {destinations} Safari", 'start_date': start_date, 'end_date': end_date, 'duration_nights': str(duration_days), 'destinations': destinations, 'travel_style': accommodation_tier.title()},
            'itinerary':    itinerary,
            'line_items':   line_items,
            'photo_cache':  photo_cache,
            'pricing':      {'total_price_usd': total_price, 'deposit_amount_usd': deposit, 'balance_amount_usd': balance, 'within_budget': itinerary_data.get('within_budget', True), 'budget_notes': itinerary_data.get('budget_notes', '')},
            'narrative':    {'intro': intro_narrative, 'days': narrative_days},
            'agent_profile':{'tagline': profile.get('tagline', 'Travel & Safari Specialists'), 'bio': profile.get('bio', ''), 'years_experience': profile.get('years_experience', ''), 'safaris_planned': profile.get('safaris_planned', ''), 'countries_covered': profile.get('countries_covered', ''), 'awards': profile.get('awards', []), 'memberships': profile.get('memberships', []), 'address': profile.get('address', ''), 'facebook': profile.get('facebook', ''), 'instagram': profile.get('instagram', ''), 'linkedin': profile.get('linkedin', '')},
            'agent_reviews': reviews_formatted,
        }

        logger.info("Generating PDF...")
        generate_quote_pdf(pdf_data, output_path)
        pdf_url = supabase_upload(output_path, filename)

        itinerary_json_payload = {
            'pricing': {
                'total_price_usd':    total_price,
                'deposit_amount_usd': deposit,
                'balance_amount_usd': balance,
            },
            'line_items': line_items,
        }

        quote_record = {
            'quote_number':             quote_number,
            'agent_id':                 agent_id,
            'status':                   'generated',
            'source':                   intake_source,
            'client_name':              client_name,
            'client_email':             client_email,
            'destinations':             destinations,
            'start_date':               start_date[:10] if start_date else None,
            'end_date':                 end_date[:10] if end_date else None,
            'duration_days':            duration_days,
            'pax_adults':               pax_adults,
            'pax_children':             pax_children,
            'accommodation_tier':       accommodation_tier,
            'total_price_usd_cents':    int(total_price * 100),
            'client_budget_usd_cents':  int(budget_usd * 100),
            'itinerary_json':           itinerary_json_payload,
            'pdf_url':                  pdf_url,
            'accept_token':             client_accept_token,
            'change_token':             client_changes_token,
            'reject_token':             reject_token,
            'special_requests':         special_requests,
        }

        try:
            insert_url = f"{SUPABASE_URL}/rest/v1/quotes"
            insert_req = urllib.request.Request(
                insert_url,
                data=json.dumps(quote_record).encode('utf-8'),
                method='POST',
                headers={
                    'Authorization': f'Bearer {SUPABASE_KEY}',
                    'apikey': SUPABASE_KEY,
                    'Content-Type': 'application/json',
                    'Prefer': 'return=minimal,resolution=merge-duplicates',
                }
            )
            with urllib.request.urlopen(insert_req, timeout=15) as resp:
                resp.read()
            logger.info(f"Quote saved to Supabase: {quote_number}, total=${total_price}")
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            logger.error(f"Quote save HTTP error {e.code}: {error_body}")
            try:
                minimal_record = {
                    'quote_number':             quote_number,
                    'agent_id':                 agent_id,
                    'status':                   'generated',
                    'source':                   intake_source,
                    'client_name':              client_name,
                    'client_email':             client_email,
                    'destinations':             destinations,
                    'start_date':               start_date[:10] if start_date else None,
                    'end_date':                 end_date[:10] if end_date else None,
                    'duration_days':            duration_days,
                    'pax_adults':               pax_adults,
                    'pax_children':             pax_children,
                    'total_price_usd_cents':    int(total_price * 100),
                    'client_budget_usd_cents':  int(budget_usd * 100),
                    'pdf_url':                  pdf_url,
                    'accept_token':             client_accept_token,
                    'change_token':             client_changes_token,
                    'reject_token':             reject_token,
                }
                insert_req2 = urllib.request.Request(
                    insert_url,
                    data=json.dumps(minimal_record).encode('utf-8'),
                    method='POST',
                    headers={
                        'Authorization': f'Bearer {SUPABASE_KEY}',
                        'apikey': SUPABASE_KEY,
                        'Content-Type': 'application/json',
                        'Prefer': 'return=minimal,resolution=merge-duplicates',
                    }
                )
                with urllib.request.urlopen(insert_req2, timeout=15) as resp:
                    resp.read()
                logger.info(f"Quote saved (minimal) to Supabase: {quote_number}")
            except urllib.error.HTTPError as e2:
                error_body2 = e2.read().decode()
                logger.error(f"Quote minimal save error {e2.code}: {error_body2}")
        except Exception as e:
            logger.error(f"Quote save error: {str(e)}")

        with open(output_path, 'rb') as f:
            pdf_bytes  = f.read()
            pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

        logger.info(f"PDF complete: {filename} ({len(pdf_bytes)} bytes)")

        agent_email_addr = agent.get('email', '')
        if agent_email_addr:
            email_html = agent_approval_email_html(
                agent_name=agent.get('agent_name', 'Agent'),
                agency_name=agent.get('agency_name', 'SafariFlow'),
                client_name=client_name,
                quote_number=quote_number,
                start_date=start_date,
                end_date=end_date,
                total_price=total_price,
                approve_url=approve_url,
                reject_url=reject_url,
                brand_primary=agent.get('brand_color_primary', '#2E4A7A'),
                brand_secondary=agent.get('brand_color_secondary', '#C4922A'),
            )
            send_email(
                to=agent_email_addr,
                subject=f"New Quote Ready for Review — {client_name} ({quote_number})",
                html=email_html,
                attachments=[{'filename': filename, 'content': pdf_base64}]
            )
            logger.info(f"Agent approval email sent to {agent_email_addr}")

        return jsonify({
            'success':           True,
            'filename':          filename,
            'quote_number':      quote_number,
            'pdf_base64':        pdf_base64,
            'pdf_url':           pdf_url,
            'file_size':         len(pdf_bytes),
            'client_name':       client_name,
            'client_email':      client_email,
            'agent_email':       agent.get('email', ''),
            'agent_name':        agent.get('agent_name', ''),
            'agency_name':       agent.get('agency_name', ''),
            'quote_date':        start_date[:10] if start_date else '',
            'total_price_usd':   total_price,
            'deposit_usd':       deposit,
            'balance_usd':       balance,
            'itinerary_days':    len(itinerary),
            'line_items':        line_items,
            'itinerary_json':    {
                'pricing': {
                    'total_price_usd':    total_price,
                    'deposit_amount_usd': deposit,
                    'balance_amount_usd': balance,
                },
                'line_items': line_items,
                'itinerary':  itinerary,
            },
            'approve_url':       approve_url,
            'reject_url':        reject_url,
            'client_accept_url': client_accept_url,
            'client_changes_url':client_changes_url,
            'destinations':      destinations,
            'start_date':        start_date,
            'end_date':          end_date,
        })

    except Exception as e:
        logger.error(f"PDF generation failed: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Helper pages ─────────────────────────────────────────────────────────────
def confirmation_page(token, action, title, message, button_label, button_color, icon):
    return f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SafariFlow</title></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f4f4f4;">
<div style="max-width:500px;margin:60px auto;background:white;padding:40px;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,0.1);text-align:center;">
    <div style="font-size:48px;margin-bottom:16px">{icon}</div>
    <h2 style="color:#1B2A47;margin-bottom:8px">{title}</h2>
    <p style="color:#444;margin-bottom:24px">{message}</p>
    <form method="POST" action="/{action}-confirm">
        <input type="hidden" name="token" value="{token}">
        <button type="submit" style="background:{button_color};color:white;border:none;padding:14px 32px;border-radius:6px;font-size:15px;font-weight:bold;cursor:pointer;margin-right:12px;">
            {button_label}
        </button>
        <a href="javascript:history.back()" style="display:inline-block;background:#ccc;color:white;padding:14px 24px;border-radius:6px;font-size:15px;font-weight:bold;text-decoration:none;">
            Cancel
        </a>
    </form>
</div>
</body></html>'''


def success_page(icon, title, message, quote_id):
    return f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SafariFlow</title></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f4f4f4;">
<div style="max-width:500px;margin:60px auto;background:white;padding:40px;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,0.1);text-align:center;">
    <div style="font-size:48px;margin-bottom:16px">{icon}</div>
    <h2 style="color:#1B2A47;margin-bottom:8px">{title}</h2>
    <p style="color:#444">{message}</p>
    <p style="color:#C4922A;font-weight:bold;font-size:14px;margin-top:16px">{quote_id}</p>
</div>
</body></html>'''


def invalid_page():
    return '''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;text-align:center;padding:60px;background:#f4f4f4;">
<div style="max-width:500px;margin:0 auto;background:white;padding:40px;border-radius:12px;">
    <h2 style="color:#c0392b">Invalid or Expired Link</h2>
    <p style="color:#444">This link is invalid or has expired. Please contact your travel specialist.</p>
</div></body></html>''', 400


# ─── Agent Approve ────────────────────────────────────────────────────────────
@app.route('/approve', methods=['GET'])
def approve():
    token = request.args.get('token', '')
    quote_id = verify_token(token, 'approve')
    if not quote_id:
        return invalid_page()
    return confirmation_page(token, 'approve',
        'Approve This Quote',
        f'You are about to approve quote <strong>{quote_id}</strong> and send it to the client. Are you sure?',
        'Yes, Approve & Send', '#1B2A47', '&#x2705;')


@app.route('/approve-confirm', methods=['POST'])
def approve_confirm():
    token = request.form.get('token', '')
    quote_id = verify_token(token, 'approve')
    if not quote_id:
        return invalid_page()

    supabase_update('quotes', {'quote_number': f'eq.{quote_id}'}, {'status': 'sent'})

    quotes = supabase_get('quotes', {'quote_number': f'eq.{quote_id}', 'select': '*'})
    if quotes:
        quote = quotes[0]
        agents = supabase_get('agents', {'id': f'eq.{quote.get("agent_id")}', 'select': '*'})
        agent = agents[0] if agents else {}

        client_email_addr    = quote.get('client_email', '')
        client_accept_token  = generate_token(quote_id, 'client-accept')
        client_changes_token = generate_token(quote_id, 'client-changes')
        accept_url  = f"{API_BASE_URL}/client-accept?token={client_accept_token}"
        changes_url = f"{API_BASE_URL}/client-changes?token={client_changes_token}"

        if client_email_addr:
            pdf_url    = quote.get('pdf_url', '')
            pdf_base64 = ''
            if pdf_url:
                try:
                    req = urllib.request.Request(pdf_url, headers={'User-Agent': 'SafariFlow/5.0'})
                    with urllib.request.urlopen(req, timeout=15) as resp:
                        pdf_base64 = base64.b64encode(resp.read()).decode('utf-8')
                except Exception as e:
                    logger.warning(f"Could not fetch PDF for attachment: {str(e)}")

            email_html = client_quote_email_html(
                client_name=quote.get('client_name', 'Dear Guest'),
                agency_name=agent.get('agency_name', 'SafariFlow'),
                agent_name=agent.get('agent_name', ''),
                agent_email=agent.get('email', ''),
                agent_phone=agent.get('phone', ''),
                quote_number=quote_id,
                start_date=str(quote.get('start_date', '')),
                end_date=str(quote.get('end_date', '')),
                accept_url=accept_url,
                changes_url=changes_url,
                brand_primary=agent.get('brand_color_primary', '#2E4A7A'),
                brand_secondary=agent.get('brand_color_secondary', '#C4922A'),
            )
            attachments = [{'filename': f'SafariFlow_Quote_{quote_id}.pdf', 'content': pdf_base64}] if pdf_base64 else []
            send_email(
                to=client_email_addr,
                subject=f"Your Safari Quote is Ready — {quote_id}",
                html=email_html,
                attachments=attachments if attachments else None,
            )
            logger.info(f"Client quote email sent to {client_email_addr}")

    trigger_make_webhook(MAKE_S2_WEBHOOK, {'event': 'quote_approved', 'quote_number': quote_id, 'approved_at': int(time.time())})
    return success_page('&#x2705;', 'Quote Approved', 'The quote has been approved and sent to the client.', quote_id)


# ─── Agent Reject ─────────────────────────────────────────────────────────────
@app.route('/reject', methods=['GET'])
def reject():
    token = request.args.get('token', '')
    quote_id = verify_token(token, 'reject')
    if not quote_id:
        return invalid_page()
    return confirmation_page(token, 'reject',
        'Request Changes',
        f'You are about to flag quote <strong>{quote_id}</strong> for revision. Are you sure?',
        'Yes, Request Changes', '#C4922A', '&#x270F;&#xFE0F;')


@app.route('/reject-confirm', methods=['POST'])
def reject_confirm():
    token = request.form.get('token', '')
    quote_id = verify_token(token, 'reject')
    if not quote_id:
        return invalid_page()
    supabase_update('quotes', {'quote_number': f'eq.{quote_id}'}, {'status': 'revision_requested'})
    trigger_make_webhook(MAKE_S2_WEBHOOK, {'event': 'quote_rejected', 'quote_number': quote_id, 'rejected_at': int(time.time())})
    return success_page('&#x270F;&#xFE0F;', 'Revision Requested', 'The quote has been flagged for revision.', quote_id)


# ─── Client Accept ────────────────────────────────────────────────────────────
@app.route('/client-accept', methods=['GET'])
def client_accept():
    token = request.args.get('token', '')
    quote_id = verify_token(token, 'client-accept')
    if not quote_id:
        return invalid_page()
    return confirmation_page(token, 'client-accept',
        'Accept This Quote',
        f'You are about to accept quote <strong>{quote_id}</strong> and confirm your safari booking. Are you sure?',
        'Yes, Accept This Quote', '#1B2A47', '&#x1F389;')


@app.route('/client-accept-confirm', methods=['POST'])
def client_accept_confirm():
    token = request.form.get('token', '')
    quote_id = verify_token(token, 'client-accept')
    if not quote_id:
        return invalid_page()

    supabase_update('quotes', {'quote_number': f'eq.{quote_id}'}, {'status': 'accepted'})
    trigger_make_webhook(MAKE_S2_WEBHOOK, {'event': 'quote_accepted', 'quote_number': quote_id, 'accepted_at': int(time.time())})
    logger.info(f"Quote accepted by client: {quote_id}")

    try:
        quotes = supabase_get('quotes', {'quote_number': f'eq.{quote_id}', 'select': '*'})
        if quotes:
            quote    = quotes[0]
            agent_id = quote.get('agent_id', '')
            if agent_id:
                inv_payload = json.dumps({'quote_id': quote_id, 'agent_id': agent_id}).encode('utf-8')
                inv_req = urllib.request.Request(
                    f"{API_BASE_URL}/generate-invoice",
                    data=inv_payload, method='POST',
                    headers={'Content-Type': 'application/json'}
                )
                with urllib.request.urlopen(inv_req, timeout=60) as resp:
                    inv_result = json.loads(resp.read().decode())
                logger.info(f"Invoice auto-generated: {inv_result.get('invoice_number')} for {quote_id}")
    except Exception as e:
        logger.error(f"Auto-invoice error for {quote_id}: {str(e)}")

    return success_page('&#x1F389;', 'Quote Accepted!',
        'Thank you! Your safari booking is confirmed. Your invoice has been sent to your email.',
        quote_id)


# ─── Client Changes ───────────────────────────────────────────────────────────
@app.route('/client-changes', methods=['GET'])
def client_changes():
    token = request.args.get('token', '')
    quote_id = verify_token(token, 'client-changes')
    if not quote_id:
        return invalid_page()

    quotes  = supabase_get('quotes', {'quote_number': f'eq.{quote_id}', 'select': '*'})
    quote   = quotes[0] if quotes else {}
    agents  = supabase_get('agents', {'id': f'eq.{quote.get("agent_id", "")}', 'select': '*'})
    agent   = agents[0] if agents else {}
    brand_primary   = agent.get('brand_color_primary', '#2E4A7A')
    brand_secondary = agent.get('brand_color_secondary', '#C4922A')
    agency_name     = agent.get('agency_name', 'SafariFlow')

    extras = supabase_get('optional_extras', {
        'agent_id': f'eq.{agent.get("id", "")}',
        'is_active': 'eq.true',
        'select': 'id,name,category,price_per_person_usd_cents,price_type,duration_hours',
        'order': 'category.asc,name.asc'
    }) if agent.get('id') else []

    extras_html = ''
    if extras:
        extras_buttons = ''
        for ex in extras:
            price = ex.get('price_per_person_usd_cents', 0) / 100
            price_label = f"${price:,.0f}/pp" if ex.get('price_type') == 'per_person' else f"${price:,.0f}/grp"
            extras_buttons += f'''
            <div class="extra-btn" onclick="toggleExtra(this)" data-id="{ex['id']}" data-name="{ex['name']}">
              <input type="hidden" name="extra_{ex['id']}" value="no" class="extra-input">
              <div class="extra-name">{ex['name']}</div>
              <div class="extra-meta">{ex.get('category','')} · {ex.get('duration_hours',2)}h · {price_label}</div>
            </div>'''
        extras_html = f'''
      <div class="section-label" style="margin-top:24px;">Would you like to add any extras?</div>
      <p style="font-size:12px;color:#888;margin-bottom:12px;">Tap to add experiences to your revised quote.</p>
      <div class="extras-grid">{extras_buttons}</div>'''

    min_date = time.strftime('%Y-%m-%d')

    return f'''<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Request Changes — {quote_id}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; background: #f4f4f4; padding: 20px; }}
  .container {{ max-width: 580px; margin: 40px auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 16px rgba(0,0,0,0.1); }}
  .header {{ background: {brand_primary}; padding: 28px; text-align: center; color: white; }}
  .header h1 {{ font-size: 20px; letter-spacing: 2px; margin-bottom: 4px; }}
  .header p {{ font-size: 12px; opacity: 0.8; }}
  .gold-line {{ background: {brand_secondary}; height: 3px; }}
  .body {{ padding: 32px; }}
  .body h2 {{ font-size: 18px; color: #1A1A1A; margin-bottom: 6px; }}
  .sub {{ font-size: 13px; color: #666; margin-bottom: 24px; }}
  .section-label {{ font-size: 11px; font-weight: bold; letter-spacing: 1px; color: #999; text-transform: uppercase; margin-bottom: 10px; }}
  .quote-ref {{ background: #F8F6F2; border-radius: 8px; padding: 12px 16px; margin-bottom: 24px; font-size: 13px; color: #666; }}
  .quote-ref span {{ color: {brand_secondary}; font-weight: bold; font-family: monospace; }}
  .changes-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 8px; }}
  .change-btn {{ display: flex; align-items: center; gap: 10px; padding: 12px 14px; border-radius: 10px; cursor: pointer; border: 2px solid #E8E4DE; background: #F8F6F2; transition: all 0.2s; user-select: none; }}
  .change-btn:hover {{ border-color: {brand_secondary}; }}
  .change-btn.selected {{ border-color: {brand_primary}; background: rgba(46,74,122,0.07); }}
  .change-btn input {{ display: none; }}
  .change-icon {{ font-size: 20px; flex-shrink: 0; }}
  .change-label {{ font-size: 13px; color: #1A1A1A; font-weight: 500; flex: 1; }}
  .change-btn.selected .change-label {{ color: {brand_primary}; font-weight: 700; }}
  .check-mark {{ width: 20px; height: 20px; border-radius: 50%; background: {brand_primary}; color: white; font-size: 11px; display: none; align-items: center; justify-content: center; flex-shrink: 0; }}
  .change-btn.selected .check-mark {{ display: flex; }}
  .expand-field {{ display: none; background: #F0EDE8; border-radius: 10px; padding: 16px; margin-top: 10px; margin-bottom: 4px; border: 1px solid #E0D8CE; }}
  .expand-field.visible {{ display: block; }}
  .expand-label {{ font-size: 11px; font-weight: bold; color: #888; text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 8px; }}
  .budget-input {{ width: 100%; padding: 10px 14px; border: 1px solid #E8E4DE; border-radius: 8px; font-size: 16px; font-weight: 600; color: #1A1A1A; outline: none; font-family: Arial, sans-serif; }}
  .budget-input:focus {{ border-color: {brand_primary}; }}
  .form-row-dates {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
  .date-input {{ width: 100%; padding: 10px 12px; border: 1px solid #E8E4DE; border-radius: 8px; font-size: 14px; font-family: Arial, sans-serif; outline: none; color: #1A1A1A; cursor: pointer; transition: border 0.2s; }}
  .date-input:focus {{ border-color: {brand_primary}; }}
  .stepper-row {{ display: flex; align-items: center; justify-content: space-between; margin-bottom: 10px; }}
  .stepper-label {{ font-size: 13px; color: #444; font-weight: 500; }}
  .stepper-controls {{ display: flex; align-items: center; gap: 12px; }}
  .stepper-btn {{ width: 32px; height: 32px; border-radius: 50%; background: {brand_primary}; color: white; border: none; font-size: 18px; cursor: pointer; display: flex; align-items: center; justify-content: center; font-weight: bold; }}
  .stepper-value {{ font-size: 16px; font-weight: 700; color: #1A1A1A; min-width: 24px; text-align: center; }}
  .extras-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-bottom: 20px; }}
  .extra-btn {{ padding: 12px 14px; border-radius: 10px; cursor: pointer; border: 2px solid #E8E4DE; background: #F8F6F2; transition: all 0.2s; user-select: none; }}
  .extra-btn.selected {{ border-color: {brand_secondary}; background: rgba(196,146,42,0.08); }}
  .extra-name {{ font-size: 13px; color: #1A1A1A; font-weight: 600; margin-bottom: 3px; }}
  .extra-btn.selected .extra-name {{ color: {brand_secondary}; }}
  .extra-meta {{ font-size: 11px; color: #888; }}
  .notes-area {{ width: 100%; padding: 12px 14px; border: 1px solid #E8E4DE; border-radius: 8px; font-size: 13px; font-family: Arial, sans-serif; outline: none; resize: vertical; min-height: 80px; transition: border 0.2s; color: #1A1A1A; }}
  .notes-area:focus {{ border-color: {brand_primary}; }}
  .submit-btn {{ width: 100%; background: {brand_primary}; color: white; border: none; padding: 14px; border-radius: 8px; font-size: 15px; font-weight: bold; cursor: pointer; letter-spacing: 1px; margin-top: 8px; }}
  @media (max-width: 480px) {{ .changes-grid, .extras-grid {{ grid-template-columns: 1fr; }} }}
</style>
</head>
<body>
<div class="container">
  <div class="header"><h1>{agency_name}</h1><p>REQUEST CHANGES TO YOUR QUOTE</p></div>
  <div class="gold-line"></div>
  <div class="body">
    <h2>What would you like changed?</h2>
    <p class="sub">Tap the options below. We'll revise your quote and send it back to you.</p>
    <div class="quote-ref">Quote Reference: <span>{quote_id}</span></div>
    <form method="POST" action="/client-changes-confirm" id="changeForm">
      <input type="hidden" name="token" value="{token}">
      <input type="hidden" name="revised_month" id="revised_month_hidden" value="">
      <input type="hidden" name="revised_adults" id="revised_adults_hidden" value="0">
      <input type="hidden" name="revised_children" id="revised_children_hidden" value="0">
      <div class="section-label">What needs changing?</div>
      <div class="changes-grid">
        <div class="change-btn" onclick="toggleChange(this, 'accommodation')"><input type="hidden" name="change_accommodation" value="no"><span class="change-icon">🏨</span><span class="change-label">Accommodation</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'dates')"><input type="hidden" name="change_dates" value="no"><span class="change-icon">📅</span><span class="change-label">Travel Dates</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'budget')"><input type="hidden" name="change_budget" value="no"><span class="change-icon">💰</span><span class="change-label">Budget</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'destinations')"><input type="hidden" name="change_destinations" value="no"><span class="change-icon">📍</span><span class="change-label">Destinations</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'travelers')"><input type="hidden" name="change_travelers" value="no"><span class="change-icon">👥</span><span class="change-label">No. of Travelers</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'transport')"><input type="hidden" name="change_transport" value="no"><span class="change-icon">✈️</span><span class="change-label">Transport</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'duration')"><input type="hidden" name="change_duration" value="no"><span class="change-icon">🌙</span><span class="change-label">Trip Duration</span><span class="check-mark">✓</span></div>
        <div class="change-btn" onclick="toggleChange(this, 'other')"><input type="hidden" name="change_other" value="no"><span class="change-icon">✏️</span><span class="change-label">Other</span><span class="check-mark">✓</span></div>
      </div>
      <div class="expand-field" id="field_budget"><div class="expand-label">Revised budget (USD)</div><input type="number" name="revised_budget" class="budget-input" placeholder="e.g. 8000" min="0"></div>
      <div class="expand-field" id="field_dates"><div class="expand-label">Select New Travel Dates</div><div class="form-row-dates"><div><div style="font-size:12px;color:#666;font-weight:600;margin-bottom:6px;">Start Date</div><input type="date" name="revised_start_date" class="date-input" min="{min_date}"></div><div><div style="font-size:12px;color:#666;font-weight:600;margin-bottom:6px;">End Date</div><input type="date" name="revised_end_date" class="date-input" min="{min_date}"></div></div></div>
      <div class="expand-field" id="field_travelers"><div class="expand-label">How many travelers?</div><div class="stepper-row"><span class="stepper-label">Adults</span><div class="stepper-controls"><button type="button" class="stepper-btn" onclick="updatePax('adults',-1)">−</button><span class="stepper-value" id="adults_display">2</span><button type="button" class="stepper-btn" onclick="updatePax('adults',1)">+</button></div></div><div class="stepper-row"><span class="stepper-label">Children</span><div class="stepper-controls"><button type="button" class="stepper-btn" onclick="updatePax('children',-1)">−</button><span class="stepper-value" id="children_display">0</span><button type="button" class="stepper-btn" onclick="updatePax('children',1)">+</button></div></div></div>
      {extras_html}
      <div style="margin-top:20px;"><div class="section-label">Anything else?</div><textarea name="notes" class="notes-area" placeholder="Any other details or requests..."></textarea></div>
      <button type="submit" class="submit-btn">✏️ SUBMIT CHANGE REQUEST</button>
    </form>
  </div>
</div>
<script>
  function toggleChange(btn, type) {{
    btn.classList.toggle('selected');
    btn.querySelector('input[type="hidden"]').value = btn.classList.contains('selected') ? 'yes' : 'no';
    var field = document.getElementById('field_' + type);
    if (field) field.classList.toggle('visible', btn.classList.contains('selected'));
  }}
  var adults = 2, children = 0;
  function updatePax(type, delta) {{
    if (type === 'adults') {{ adults = Math.max(1, adults + delta); document.getElementById('adults_display').textContent = adults; document.getElementById('revised_adults_hidden').value = adults; }}
    else {{ children = Math.max(0, children + delta); document.getElementById('children_display').textContent = children; document.getElementById('revised_children_hidden').value = children; }}
  }}
  function toggleExtra(btn) {{
    btn.classList.toggle('selected');
    btn.querySelector('.extra-input').value = btn.classList.contains('selected') ? 'yes' : 'no';
  }}
</script>
</body></html>'''


@app.route('/client-changes-confirm', methods=['POST'])
def client_changes_confirm():
    token = request.form.get('token', '')
    quote_id = verify_token(token, 'client-changes')
    if not quote_id:
        return invalid_page()

    selected_extras = [key.replace('extra_', '') for key, value in request.form.items() if key.startswith('extra_') and value == 'yes']

    change_request = {
        'accommodation':     request.form.get('change_accommodation') == 'yes',
        'dates':             request.form.get('change_dates') == 'yes',
        'budget':            request.form.get('change_budget') == 'yes',
        'destinations':      request.form.get('change_destinations') == 'yes',
        'travelers':         request.form.get('change_travelers') == 'yes',
        'transport':         request.form.get('change_transport') == 'yes',
        'duration':          request.form.get('change_duration') == 'yes',
        'other':             request.form.get('change_other') == 'yes',
        'revised_budget':    request.form.get('revised_budget', ''),
        'revised_start_date':request.form.get('revised_start_date', ''),
        'revised_end_date':  request.form.get('revised_end_date', ''),
        'revised_month':     request.form.get('revised_month', ''),
        'revised_adults':    request.form.get('revised_adults', '0'),
        'revised_children':  request.form.get('revised_children', '0'),
        'preferred_month':   request.form.get('revised_month', ''),
        'notes':             request.form.get('notes', ''),
        'selected_extras':   selected_extras,
    }

    changes_list = []
    if change_request['accommodation']: changes_list.append('Accommodation')
    if change_request['dates']:         changes_list.append('Travel Dates')
    if change_request['budget']:        changes_list.append('Budget')
    if change_request['destinations']:  changes_list.append('Destinations')
    if change_request['travelers']:     changes_list.append('Number of Travelers')
    if change_request['transport']:     changes_list.append('Transport')
    if change_request['duration']:      changes_list.append('Trip Duration')
    if change_request['other']:         changes_list.append('Other')

    supabase_update('quotes', {'quote_number': f'eq.{quote_id}'}, {
        'status': 'revision_requested',
        'change_request': json.dumps(change_request),
    })

    trigger_make_webhook(MAKE_S3_WEBHOOK, {
        'event':             'client_changes_requested',
        'quote_number':      quote_id,
        'changes_requested': changes_list,
        'revised_budget':    change_request['revised_budget'],
        'preferred_month':   change_request['preferred_month'],
        'notes':             change_request['notes'],
        'requested_at':      int(time.time()),
    })

    quotes = supabase_get('quotes', {'quote_number': f'eq.{quote_id}', 'select': '*'})
    if quotes:
        quote  = quotes[0]
        agents = supabase_get('agents', {'id': f'eq.{quote.get("agent_id")}', 'select': '*'})
        agent  = agents[0] if agents else {}
        agent_email_addr = agent.get('email', '')
        if agent_email_addr:
            brand_primary   = agent.get('brand_color_primary', '#2E4A7A')
            brand_secondary = agent.get('brand_color_secondary', '#C4922A')
            agency_name     = agent.get('agency_name', 'SafariFlow')
            agent_name      = agent.get('agent_name', 'Agent')
            portal_link     = f"{PORTAL_URL}/quotes/review/{quote_id}"
            changes_html    = ''.join([f'<li style="margin-bottom:6px;color:#1A1A1A;font-size:13px;">{c}</li>' for c in changes_list])
            budget_line     = f'<p style="margin:12px 0;font-size:13px;color:#1A1A1A;"><strong>Revised budget:</strong> ${float(change_request["revised_budget"]):,.0f}</p>' if change_request.get('revised_budget') else ''
            notes_line      = f'<div style="background:#F8F6F2;border-radius:8px;padding:12px 14px;margin:12px 0;font-size:13px;color:#444;">💬 {change_request["notes"]}</div>' if change_request.get('notes') else ''

            revision_email_html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;font-family:Arial,sans-serif;background:#f4f4f4;">
<table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center" style="padding:30px 0;">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.1);">
<tr><td style="background-color:{brand_primary};padding:30px;text-align:center;">
  <h1 style="margin:0;color:#FFFFFF;font-size:22px;letter-spacing:2px;">{agency_name}</h1>
  <p style="margin:6px 0 0;color:#FFFFFF;font-size:12px;letter-spacing:1px;">CLIENT CHANGE REQUEST</p>
</td></tr>
<tr><td style="background-color:{brand_secondary};height:3px;"></td></tr>
<tr><td style="padding:36px 40px;background:#ffffff;">
  <p style="color:#1A1A1A;font-size:16px;margin:0 0 16px;">Hello <strong>{agent_name}</strong>, your client has requested changes to quote <strong>{quote_id}</strong>.</p>
  <div style="background:#FFF8F0;border:1px solid #F5DFB0;border-radius:8px;padding:16px 20px;margin-bottom:20px;">
    <p style="font-size:12px;font-weight:bold;color:#888;letter-spacing:1px;text-transform:uppercase;margin-bottom:10px;">Changes Requested:</p>
    <ul style="margin:0;padding-left:18px;">{changes_html}</ul>{budget_line}{notes_line}
  </div>
  <table width="100%" cellpadding="0" cellspacing="0" style="margin:16px 0;">
    <tr><td align="center"><a href="{portal_link}" style="display:inline-block;background-color:{brand_primary};color:#FFFFFF;padding:16px 32px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:14px;letter-spacing:1px;">🔍 REVIEW REVISED QUOTE</a></td></tr>
  </table>
</td></tr>
<tr><td style="background:#F5F0E8;padding:20px 40px;text-align:center;">
  <p style="margin:0;color:#444444;font-size:12px;">{agency_name} · Powered by SafariFlow</p>
</td></tr>
</table></td></tr></table>
</body></html>"""

            send_email(to=agent_email_addr, subject=f"Client Requested Changes — {quote_id}", html=revision_email_html)

    return success_page('✏️', 'Changes Received!',
        'Thank you! We have received your change request and will send you a revised quote shortly.',
        quote_id)


# ─── Package Safari PDF ───────────────────────────────────────────────────────
@app.route('/generate-package-pdf', methods=['POST'])
def generate_package_pdf():
    try:
        data             = request.get_json(force=True)
        agent_id         = data.get('agent_id', '')
        package_id       = data.get('package_id', '')
        client_name      = data.get('client_name', '')
        client_email     = data.get('client_email', '')
        client_phone     = data.get('client_phone', '')
        client_nationality = data.get('client_nationality', '')
        start_date       = data.get('start_date', '')
        end_date         = data.get('end_date', '')
        pax_adults       = int(data.get('pax_adults', 2))
        pax_children     = int(data.get('pax_children', 0))
        special_requests = data.get('special_requests', [])
        request_id       = data.get('request_id', str(uuid.uuid4())[:8])

        agents = supabase_get('agents', {'id': f'eq.{agent_id}', 'select': '*'})
        if not agents:
            return jsonify({'error': 'Agent not found'}), 404
        agent = agents[0]

        packages = supabase_get('packages', {'id': f'eq.{package_id}', 'select': '*'})
        if not packages:
            return jsonify({'error': 'Package not found'}), 404
        pkg = packages[0]

        base_price       = float(pkg.get('base_price_usd_cents', 0)) / 100
        markup_pct       = float(agent.get('markup_overall_pct', 0) or 0) / 100
        total_per_person = round(base_price * (1 + markup_pct), 2)
        total_price      = round(total_per_person * pax_adults, 2)
        deposit_pct      = float(agent.get('deposit_percentage', 30) or 30) / 100
        deposit          = round(total_price * deposit_pct, 2)
        balance          = round(total_price - deposit, 2)
        duration_days    = pkg.get('duration_days', 7)
        destinations     = pkg.get('destination', '')

        pkg_itinerary = pkg.get('itinerary_days', [])
        if not pkg_itinerary:
            pkg_itinerary = [{'day_number': i+1, 'destination': destinations, 'title': f'Day {i+1}', 'accommodation_name': pkg.get('accommodation_name', ''), 'room_type': 'Standard Room', 'meal_plan': 'Full Board', 'nights': 1, 'transport_description': None, 'image_search_query': f'{destinations} safari wildlife'} for i in range(duration_days)]

        line_items = [{'line_type': 'accommodation', 'description': pkg.get('name', ''), 'details': f'{duration_days} nights · {destinations}', 'quantity': pax_adults, 'unit_price': total_per_person, 'total_price': total_price, 'cost_unit_price': base_price, 'cost_total_price': round(base_price * pax_adults, 2), 'markup_pct': round(markup_pct * 100, 2), 'profit': round(total_price - base_price * pax_adults, 2)}]

        try:
            narrative_result = call_claude(f"Write a warm 3-sentence safari intro for {client_name} booking {pkg.get('name','')} for {duration_days} days in {destinations}. Return only the text.", max_tokens=300)
            intro_narrative = narrative_result if isinstance(narrative_result, str) else f"{client_name.split()[0]}, your {duration_days}-day safari awaits."
        except Exception:
            intro_narrative = f"{client_name.split()[0]}, your {duration_days}-day safari awaits."

        narrative_days = [{'day_number': d.get('day_number'), 'narrative': d.get('narrative', f"Explore {d.get('destination', destinations)}."), 'highlight': d.get('highlight', f"Wildlife in {d.get('destination', destinations)}"), 'accommodation_description': f"{d.get('accommodation_name', '')}, {d.get('room_type', 'Standard Room')} — your base."} for d in pkg_itinerary]

        quote_number         = f"QT-PKG-{request_id}"
        approve_token        = generate_token(quote_number, 'approve')
        reject_token         = generate_token(quote_number, 'reject')
        client_accept_token  = generate_token(quote_number, 'client-accept')
        client_changes_token = generate_token(quote_number, 'client-changes')
        approve_url          = f"{API_BASE_URL}/approve?token={approve_token}"
        reject_url           = f"{API_BASE_URL}/reject?token={reject_token}"
        client_accept_url    = f"{API_BASE_URL}/client-accept?token={client_accept_token}"
        client_changes_url   = f"{API_BASE_URL}/client-changes?token={client_changes_token}"

        photo_cache = fetch_photos_for_itinerary(pkg_itinerary)
        filename    = f"SafariFlow_Package_{quote_number}.pdf"
        output_path = os.path.join(OUTPUT_DIR, filename)

        pdf_data = {
            'quote_id': quote_number, 'generated_at': start_date[:10] if start_date else '',
            'accept_url': '#', 'changes_url': '#',
            'inclusions': pkg.get('inclusions', '- All accommodation\n- All meals\n- All game drives\n- Park fees'),
            'exclusions': pkg.get('exclusions', '- International flights\n- Travel insurance\n- Personal expenses'),
            'terms': agent.get('cancellation_terms') or 'This quote is valid for 14 days.',
            'agent': {'name': agent.get('agent_name', ''), 'email': agent.get('email', ''), 'phone': agent.get('phone', ''), 'agency': agent.get('agency_name', ''), 'logo_url': agent.get('logo_url', ''), 'website': agent.get('website', '')},
            'client': {'name': client_name, 'email': client_email, 'phone': client_phone, 'pax_adults': str(pax_adults), 'pax_children': str(pax_children), 'nationality': client_nationality},
            'trip': {'title': pkg.get('name', ''), 'start_date': start_date, 'end_date': end_date, 'duration_nights': str(duration_days), 'destinations': destinations, 'travel_style': pkg.get('category', 'Safari')},
            'itinerary': pkg_itinerary, 'line_items': line_items, 'photo_cache': photo_cache,
            'pricing': {'total_price_usd': total_price, 'deposit_amount_usd': deposit, 'balance_amount_usd': balance, 'within_budget': True, 'budget_notes': ''},
            'narrative': {'intro': intro_narrative, 'days': narrative_days},
            'agent_profile': {'tagline': 'Travel & Safari Specialists', 'bio': '', 'years_experience': '', 'safaris_planned': '', 'countries_covered': '', 'awards': [], 'memberships': [], 'address': '', 'facebook': '', 'instagram': '', 'linkedin': ''},
            'agent_reviews': [],
        }

        generate_quote_pdf(pdf_data, output_path)
        pdf_url = supabase_upload(output_path, filename)

        with open(output_path, 'rb') as f:
            pdf_bytes  = f.read()
            pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

        agent_email_addr = agent.get('email', '')
        if agent_email_addr:
            email_html = agent_approval_email_html(
                agent_name=agent.get('agent_name', 'Agent'), agency_name=agent.get('agency_name', 'SafariFlow'),
                client_name=client_name, quote_number=quote_number, start_date=start_date, end_date=end_date,
                total_price=total_price, approve_url=approve_url, reject_url=reject_url,
                brand_primary=agent.get('brand_color_primary', '#2E4A7A'), brand_secondary=agent.get('brand_color_secondary', '#C4922A'),
            )
            send_email(to=agent_email_addr, subject=f"New Package Quote Ready — {client_name} ({quote_number})", html=email_html, attachments=[{'filename': filename, 'content': pdf_base64}])

        return jsonify({'success': True, 'filename': filename, 'quote_number': quote_number, 'pdf_base64': pdf_base64, 'pdf_url': pdf_url, 'client_name': client_name, 'client_email': client_email, 'agent_email': agent.get('email', ''), 'total_price_usd': total_price, 'deposit_usd': deposit, 'balance_usd': balance, 'approve_url': approve_url, 'reject_url': reject_url, 'client_accept_url': client_accept_url, 'client_changes_url': client_changes_url})

    except Exception as e:
        logger.error(f"Package PDF failed: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Invoicing ────────────────────────────────────────────────────────────────
def generate_invoice_number(agent_id):
    import datetime
    year = datetime.date.today().year
    existing = supabase_get('invoices', {'agent_id': f'eq.{agent_id}', 'invoice_number': f'like.INV-{year}-%', 'select': 'invoice_number', 'order': 'created_at.desc', 'limit': '1'})
    if existing:
        try:
            last_num = int(existing[0]['invoice_number'].split('-')[-1])
            return f"INV-{year}-{str(last_num + 1).zfill(4)}"
        except Exception:
            pass
    return f"INV-{year}-0001"


@app.route('/generate-invoice', methods=['POST'])
def generate_invoice():
    try:
        data     = request.get_json(force=True)
        quote_id = data.get('quote_id', '')
        agent_id = data.get('agent_id', '')

        if not quote_id or not agent_id:
            return jsonify({'error': 'quote_id and agent_id required'}), 400

        quotes = supabase_get('quotes', {'quote_number': f'eq.{quote_id}', 'select': '*'})
        if not quotes:
            return jsonify({'error': 'Quote not found'}), 404
        quote = quotes[0]

        agents = supabase_get('agents', {'id': f'eq.{agent_id}', 'select': '*'})
        if not agents:
            return jsonify({'error': 'Agent not found'}), 404
        agent = agents[0]

        import datetime
        today        = datetime.date.today()
        deposit_pct  = float(agent.get('deposit_percentage', 30) or 30)
        balance_days = int(agent.get('balance_due_days', 60) or 60)
        total_cents  = int(quote.get('total_price_usd_cents', 0) or 0)

        if total_cents == 0:
            try:
                itinerary = quote.get('itinerary_json')
                if itinerary:
                    if isinstance(itinerary, str):
                        itinerary = json.loads(itinerary)
                    pricing   = itinerary.get('pricing', {})
                    total_usd = float(pricing.get('total_price_usd', 0) or 0)
                    if total_usd == 0:
                        total_usd = sum(float(i.get('total_price', 0)) for i in itinerary.get('line_items', []))
                    total_cents = int(total_usd * 100)
            except Exception as e:
                logger.warning(f"Could not read total from itinerary_json: {e}")

        deposit_cents = int(round(total_cents * deposit_pct / 100))
        balance_cents = total_cents - deposit_cents
        deposit_due   = str(today + datetime.timedelta(days=7))
        start_date    = quote.get('start_date', '')
        if start_date:
            try:
                sd          = datetime.date.fromisoformat(str(start_date)[:10])
                balance_due = str(sd - datetime.timedelta(days=balance_days))
            except Exception:
                balance_due = str(today + datetime.timedelta(days=30))
        else:
            balance_due = str(today + datetime.timedelta(days=30))

        inv_number = generate_invoice_number(agent_id)

        line_items = []
        try:
            itinerary_json = quote.get('itinerary_json')
            if itinerary_json:
                if isinstance(itinerary_json, str):
                    itinerary_json = json.loads(itinerary_json)
                line_items = itinerary_json.get('line_items', [])
        except Exception:
            pass

        if not line_items:
            line_items = [{'description': f"Safari — {quote.get('destinations', '')}", 'details': f"{quote.get('duration_days', '')} nights · {quote.get('pax_adults', 2)} adults", 'quantity': quote.get('pax_adults', 2), 'unit_price': round(total_cents / 100 / max(int(quote.get('pax_adults', 2)), 1), 2), 'total_price': round(total_cents / 100, 2)}]

        invoice_id     = str(uuid.uuid4())
        invoice_record = {
            'id':                    invoice_id,
            'invoice_number':        inv_number,
            'quote_id':              quote_id,
            'agent_id':              agent_id,
            'client_name':           quote.get('client_name', ''),
            'client_email':          quote.get('client_email', ''),
            'client_phone':          quote.get('client_phone', ''),
            'client_nationality':    quote.get('client_nationality', ''),
            'destinations':          quote.get('destinations', ''),
            'start_date':            quote.get('start_date', None),
            'end_date':              quote.get('end_date', None),
            'pax_adults':            int(quote.get('pax_adults', 2) or 2),
            'pax_children':          int(quote.get('pax_children', 0) or 0),
            'duration_nights':       int(quote.get('duration_days', 0) or 0),
            'subtotal_usd_cents':    total_cents,
            'tax_usd_cents':         0,
            'total_usd_cents':       total_cents,
            'deposit_pct':           deposit_pct,
            'deposit_usd_cents':     deposit_cents,
            'balance_usd_cents':     balance_cents,
            'amount_paid_usd_cents': 0,
            'amount_due_usd_cents':  total_cents,
            'deposit_due_date':      deposit_due,
            'balance_due_date':      balance_due,
            'status':                'sent',
            'line_items':            json.dumps(line_items),
        }

        insert_url = f"{SUPABASE_URL}/rest/v1/invoices"
        req = urllib.request.Request(insert_url, data=json.dumps(invoice_record).encode('utf-8'), method='POST',
            headers={'Authorization': f'Bearer {SUPABASE_KEY}', 'apikey': SUPABASE_KEY, 'Content-Type': 'application/json', 'Prefer': 'return=minimal'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            resp.read()

        from pdf_generator import generate_invoice_pdf
        pdf_data = {
            'agent': {'agency': agent.get('agency_name', ''), 'email': agent.get('email', ''), 'phone': agent.get('phone', ''), 'logo_url': agent.get('logo_url', ''), 'brand_color_primary': agent.get('brand_color_primary', '#2E4A7A'), 'brand_color_secondary': agent.get('brand_color_secondary', '#C4922A'), 'cancellation_terms': agent.get('cancellation_terms', ''), 'amendment_terms': agent.get('amendment_terms', ''), 'bank_details': agent.get('bank_details', ''), 'mpesa_details': agent.get('mpesa_details', '')},
            'client': {'name': quote.get('client_name', ''), 'email': quote.get('client_email', ''), 'phone': quote.get('client_phone', '')},
            'invoice': {'invoice_number': inv_number, 'quote_id': quote_id, 'issued_at': str(today), 'total_usd_cents': total_cents, 'deposit_usd_cents': deposit_cents, 'balance_usd_cents': balance_cents, 'deposit_due_date': deposit_due, 'balance_due_date': balance_due, 'destinations': quote.get('destinations', ''), 'start_date': str(quote.get('start_date', ''))[:10], 'end_date': str(quote.get('end_date', ''))[:10], 'pax_adults': int(quote.get('pax_adults', 2) or 2), 'pax_children': int(quote.get('pax_children', 0) or 0)},
            'line_items': line_items,
        }

        filename    = f"SafariFlow_Invoice_{inv_number}.pdf"
        output_path = os.path.join(OUTPUT_DIR, filename)
        generate_invoice_pdf(pdf_data, output_path)
        pdf_url = supabase_upload(output_path, filename)
        supabase_update('invoices', {'id': f'eq.{invoice_id}'}, {'pdf_url': pdf_url})

        with open(output_path, 'rb') as f:
            pdf_base64 = base64.b64encode(f.read()).decode('utf-8')

        client_email_addr = quote.get('client_email', '')
        if client_email_addr:
            invoice_html = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
              <div style="background:{agent.get('brand_color_primary','#2E4A7A')};padding:28px;text-align:center;color:white;border-radius:12px 12px 0 0">
                <h1 style="margin:0;font-size:20px;letter-spacing:2px">{agent.get('agency_name','')}</h1>
                <p style="margin:8px 0 0;opacity:0.8;font-size:12px">INVOICE</p>
              </div>
              <div style="background:{agent.get('brand_color_secondary','#C4922A')};height:3px"></div>
              <div style="padding:28px;background:#ffffff">
                <h2 style="color:#1A1A1A">Dear {quote.get('client_name','')},</h2>
                <p style="color:#444;line-height:1.7">Please find attached your invoice <strong>{inv_number}</strong> for your upcoming safari.</p>
                <div style="background:#F8F6F2;border-radius:8px;padding:16px;margin:20px 0">
                  <p style="margin:0 0 8px;font-weight:bold">Payment Summary:</p>
                  <p style="margin:4px 0;color:#444;font-size:13px">Total: <strong>${total_cents/100:,.2f}</strong></p>
                  <p style="margin:4px 0;color:#444;font-size:13px">Deposit due by {deposit_due}: <strong>${deposit_cents/100:,.2f}</strong></p>
                  <p style="margin:4px 0;color:#444;font-size:13px">Balance due by {balance_due}: <strong>${balance_cents/100:,.2f}</strong></p>
                </div>
                <p style="color:#444;font-size:13px">Kind regards,<br/><strong>{agent.get('agent_name','')}</strong><br/>{agent.get('agency_name','')}</p>
              </div>
            </div>"""
            send_email(to=client_email_addr, subject=f"Invoice {inv_number} — {agent.get('agency_name','')}", html=invoice_html, attachments=[{'filename': filename, 'content': pdf_base64}])

        agent_email_addr = agent.get('email', '')
        if agent_email_addr:
            send_email(to=agent_email_addr, subject=f"Invoice Sent — {quote.get('client_name','')} ({inv_number})",
                html=f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:28px">
                  <h2>Invoice Sent ✅</h2>
                  <p>Invoice <strong>{inv_number}</strong> has been sent to {quote.get('client_name','')}.</p>
                  <p>Total: <strong>${total_cents/100:,.2f}</strong><br/>Deposit due: {deposit_due}<br/>Balance due: {balance_due}</p>
                </div>""")

        logger.info(f"Invoice generated: {inv_number} for {quote_id}")
        return jsonify({'success': True, 'invoice_id': invoice_id, 'invoice_number': inv_number, 'pdf_url': pdf_url, 'total': total_cents/100, 'deposit': deposit_cents/100, 'balance': balance_cents/100, 'deposit_due': deposit_due, 'balance_due': balance_due})

    except Exception as e:
        logger.error(f"Invoice generation error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Confirm Payment — FIXED: uses agent_id from invoice, not request ─────────
@app.route('/confirm-payment', methods=['POST'])
def confirm_payment():
    """Agent confirms a payment received."""
    try:
        data           = request.get_json(force=True)
        invoice_id     = data.get('invoice_id', '')
        agent_id       = data.get('agent_id', '')  # kept for logging only
        payment_type   = data.get('payment_type', 'deposit')
        payment_method = data.get('payment_method', 'bank_transfer')
        amount_usd     = float(data.get('amount_usd', 0))
        reference      = data.get('reference', '')
        notes          = data.get('notes', '')

        logger.info(f"Payment confirm request: invoice={invoice_id}, agent={agent_id}, amount=${amount_usd}, type={payment_type}")

        if not invoice_id:
            return jsonify({'error': 'invoice_id required'}), 400

        # Always use agent_id from the invoice record itself —
        # fixes mismatch between test agent (000...001) and real Clerk agent
        invoices = supabase_get('invoices', {'id': f'eq.{invoice_id}', 'select': '*'})
        if not invoices:
            return jsonify({'error': 'Invoice not found'}), 404
        invoice = invoices[0]

        invoice_agent_id = invoice.get('agent_id', agent_id)
        logger.info(f"Using invoice agent_id: {invoice_agent_id} (request sent: {agent_id})")

        amount_cents = int(amount_usd * 100)

        payment_id     = str(uuid.uuid4())
        payment_record = {
            'id':               payment_id,
            'invoice_id':       invoice_id,
            'agent_id':         invoice_agent_id,
            'payment_type':     payment_type,
            'payment_method':   payment_method,
            'amount_usd_cents': amount_cents,
            'reference':        reference,
            'notes':            notes,
            'confirmed_by':     'agent',
        }
        insert_url = f"{SUPABASE_URL}/rest/v1/payments"
        req = urllib.request.Request(insert_url, data=json.dumps(payment_record).encode('utf-8'), method='POST',
            headers={'Authorization': f'Bearer {SUPABASE_KEY}', 'apikey': SUPABASE_KEY, 'Content-Type': 'application/json', 'Prefer': 'return=minimal'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            resp.read()
        logger.info(f"Payment record saved: {payment_id}, amount=${amount_usd}")

        paid_so_far   = int(invoice.get('amount_paid_usd_cents', 0) or 0) + amount_cents
        total_cents   = int(invoice.get('total_usd_cents', 0) or 0)
        amount_due    = max(0, total_cents - paid_so_far)
        deposit_cents = int(invoice.get('deposit_usd_cents', 0) or 0)

        if paid_so_far >= total_cents:
            new_status = 'balance_paid'
        elif paid_so_far >= deposit_cents:
            new_status = 'deposit_paid'
        else:
            new_status = invoice.get('status', 'sent')

        supabase_update('invoices', {'id': f'eq.{invoice_id}'}, {
            'amount_paid_usd_cents': paid_so_far,
            'amount_due_usd_cents':  amount_due,
            'status':                new_status,
        })

        quote_id = invoice.get('quote_id', '')
        if quote_id and new_status in ('deposit_paid', 'balance_paid'):
            supabase_update('quotes', {'quote_number': f'eq.{quote_id}'}, {
                'status': 'confirmed' if new_status == 'balance_paid' else 'deposit_paid'
            })

        logger.info(f"Payment confirmed: {payment_type} ${amount_usd} for invoice {invoice_id} → {new_status}")
        return jsonify({'success': True, 'payment_id': payment_id, 'new_status': new_status, 'amount_paid': paid_so_far/100, 'amount_due': amount_due/100})

    except Exception as e:
        logger.error(f"Payment confirmation error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Upload Inventory ─────────────────────────────────────────────────────────
@app.route('/upload-inventory/<inventory_type>', methods=['POST'])
def upload_inventory(inventory_type):
    try:
        import openpyxl

        agent_id = request.form.get('agent_id', '')
        if not agent_id:
            return jsonify({'error': 'agent_id required'}), 400
        if inventory_type not in ['accommodations', 'transport', 'park_fees']:
            return jsonify({'error': 'Invalid inventory type'}), 400
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload an Excel file (.xlsx)'}), 400

        temp_path = os.path.join(OUTPUT_DIR, f'temp_{agent_id}_{inventory_type}.xlsx')
        file.save(temp_path)

        wb = openpyxl.load_workbook(temp_path, data_only=True)
        ws = wb.active

        rows_inserted = 0
        rows_skipped  = 0
        errors        = []

        for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            if not any(row):
                continue
            first_val = str(row[0] or '').strip()
            if not first_val:
                continue

            try:
                if inventory_type == 'accommodations':
                    record = {'id': str(uuid.uuid4()), 'agent_id': agent_id, 'name': str(row[0] or '').strip(), 'destination': str(row[1] or '').strip(), 'category': str(row[2] or '').strip(), 'room_type': str(row[3] or '').strip(), 'meal_plan': str(row[4] or '').strip(), 'price_per_person_usd_cents': int(float(row[5] or 0) * 100), 'child_price_per_person_usd_cents': int(float(row[6] or 0) * 100) if row[6] else None, 'child_age_min': int(row[7]) if row[7] else 2, 'child_age_max': int(row[8]) if row[8] else 12, 'notes': str(row[9] or '').strip() or None}
                    if not record['name'] or not record['destination']:
                        rows_skipped += 1; continue
                    table = 'accommodations'

                elif inventory_type == 'transport':
                    record = {'id': str(uuid.uuid4()), 'agent_id': agent_id, 'transport_mode': str(row[0] or 'Road').strip(), 'from_location': str(row[1] or '').strip(), 'to_location': str(row[2] or '').strip(), 'operator_name': str(row[3] or '').strip(), 'transport_type': str(row[3] or '').strip(), 'pricing_type': str(row[4] or 'per_vehicle_per_day').strip(), 'price_per_person_usd_cents': int(float(row[5] or 0) * 100), 'child_price_per_person_usd_cents': int(float(row[6] or 0) * 100) if row[6] else None, 'child_age_min': int(row[7]) if row[7] else 2, 'child_age_max': int(row[8]) if row[8] else 12, 'duration_hours': float(row[9]) if row[9] else None, 'max_passengers': int(row[10]) if row[10] else None, 'notes': str(row[11] or '').strip() or None}
                    if not record['from_location'] or not record['to_location']:
                        rows_skipped += 1; continue
                    table = 'transport_routes'

                elif inventory_type == 'park_fees':
                    record = {'id': str(uuid.uuid4()), 'agent_id': agent_id, 'park_name': str(row[0] or '').strip(), 'destination': str(row[1] or '').strip(), 'visitor_category': str(row[2] or 'Non-Resident').strip(), 'fee_per_person_per_day_usd_cents': int(float(row[3] or 0) * 100), 'child_fee_per_person_per_day_usd_cents': int(float(row[4] or 0) * 100) if row[4] else None, 'child_age_min': int(row[5]) if row[5] else 3, 'child_age_max': int(row[6]) if row[6] else 17, 'notes': str(row[7] or '').strip() or None}
                    if not record['park_name']:
                        rows_skipped += 1; continue
                    table = 'park_fees'

                insert_url = f"{SUPABASE_URL}/rest/v1/{table}"
                req = urllib.request.Request(insert_url, data=json.dumps(record).encode('utf-8'), method='POST',
                    headers={'Authorization': f'Bearer {SUPABASE_KEY}', 'apikey': SUPABASE_KEY, 'Content-Type': 'application/json', 'Prefer': 'return=minimal'})
                with urllib.request.urlopen(req, timeout=15) as resp:
                    resp.read()
                rows_inserted += 1

            except Exception as row_err:
                errors.append(f"Row {row_num}: {str(row_err)}")
                rows_skipped += 1

        try:
            os.remove(temp_path)
        except Exception:
            pass

        return jsonify({'success': True, 'inventory_type': inventory_type, 'rows_inserted': rows_inserted, 'rows_skipped': rows_skipped, 'errors': errors[:5]})

    except Exception as e:
        logger.error(f"Inventory upload error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Template Downloads ───────────────────────────────────────────────────────
@app.route('/download-template/<template_type>', methods=['GET'])
def download_template(template_type):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        NAVY = '2E4A7A'; GOLD = 'C4922A'; EXAMPLE = 'FFF8EC'; BC = 'CCCCCC'

        def side(): return Side(style='thin', color=BC)
        def border(): return Border(left=side(), right=side(), top=side(), bottom=side())

        def hdr(ws, row, col, val, width=20):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=NAVY)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border()
            ws.column_dimensions[get_column_letter(col)].width = width

        def ex(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name='Arial', size=9, italic=True, color='666666')
            c.fill = PatternFill('solid', start_color=EXAMPLE)
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = border()

        def title(ws, t, sub, ncols):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            c = ws.cell(row=1, column=1, value=t)
            c.font = Font(bold=True, name='Arial', size=14, color='FFFFFF')
            c.fill = PatternFill('solid', start_color=NAVY)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 28
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            c2 = ws.cell(row=2, column=1, value=sub)
            c2.font = Font(name='Arial', size=10, color=GOLD)
            c2.fill = PatternFill('solid', start_color='F8F6F2')
            c2.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 20
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
            c3 = ws.cell(row=3, column=1, value='ROW 5 IS AN EXAMPLE — Delete before uploading. Fill from row 6 onwards.')
            c3.font = Font(bold=True, name='Arial', size=9, color='CC0000')
            c3.fill = PatternFill('solid', start_color='FFF0F0')
            c3.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[3].height = 16

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.freeze_panes = 'A5'

        if template_type == 'accommodations':
            ws.title = 'Accommodations'
            cols = [('Property Name',22),('Destination',18),('Category',18),('Room Type',18),('Meal Plan',16),('Adult Price Per Person (USD)',22),('Child Price Per Person (USD)',22),('Child Age Min',14),('Child Age Max',14),('Notes',28)]
            title(ws, 'SafariFlow — Accommodations Template', 'One row per property / room type combination.', len(cols))
            for i, (h, w) in enumerate(cols, 1): hdr(ws, 4, i, h, w)
            for i, v in enumerate(['Angama Mara','Masai Mara','Luxury','Tent Suite','Full Board',850,425,6,12,'Min 2 nights'], 1): ex(ws, 5, i, v)
            filename = 'SafariFlow_Accommodations_Template.xlsx'

        elif template_type == 'transport':
            ws.title = 'Transport'
            cols = [('Transport Mode',16),('From Location',20),('To Location',20),('Operator / Vehicle Name',24),('Pricing Type',26),('Price (USD)',16),('Child Price (USD)',18),('Child Age Min',14),('Child Age Max',14),('Duration (Hours)',16),('Max Passengers',16),('Notes',28)]
            title(ws, 'SafariFlow — Transport Template', 'Road = per_vehicle_per_day · Flight = per_person_per_sector · Train = per_person_per_trip', len(cols))
            for i, (h, w) in enumerate(cols, 1): hdr(ws, 4, i, h, w)
            for r, row in enumerate([['Road','Nairobi','Masai Mara','Land Cruiser 4x4','per_vehicle_per_day',200,'',2,12,6,7,'Pop-up roof'],['Flight','Nairobi (WIL)','Masai Mara (MRE)','Safarilink','per_person_per_sector',150,115,2,12,0.5,12,'Morning only'],['Train','Nairobi','Mombasa','SGR Economy','per_person_per_trip',30,15,2,12,4.5,'','Book 2wks ahead']], 5):
                for c, v in enumerate(row, 1): ex(ws, r, c, v)
            filename = 'SafariFlow_Transport_Template.xlsx'

        elif template_type == 'park_fees':
            ws.title = 'Park Fees'
            cols = [('Park / Reserve Name',26),('Destination / Region',22),('Visitor Category',24),('Fee Per Person Per Day (USD)',24),('Child Fee Per Day (USD)',22),('Child Age Min',14),('Child Age Max',14),('Notes',30)]
            title(ws, 'SafariFlow — Park Fees Template', 'One row per park per visitor category.', len(cols))
            for i, (h, w) in enumerate(cols, 1): hdr(ws, 4, i, h, w)
            for r, row in enumerate([['Masai Mara National Reserve','Masai Mara','Non-Resident',80,40,3,17,'KWS 2025'],['Masai Mara National Reserve','Masai Mara','East Africa Resident',35,20,3,17,'EAC passport required'],['Amboseli National Park','Amboseli','Non-Resident',60,30,3,17,'KWS 2025'],['Serengeti National Park','Serengeti','Non-Resident',70,35,5,17,'TANAPA 2025']], 5):
                for c, v in enumerate(row, 1): ex(ws, r, c, v)
            filename = 'SafariFlow_ParkFees_Template.xlsx'
        else:
            return jsonify({'error': 'Invalid template type'}), 400

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Template download error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── Health ───────────────────────────────────────────────────────────────────
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'SafariFlow API v5'})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
